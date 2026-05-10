from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
import time

options = Options()
options.add_argument("--headless")
driver = webdriver.Chrome(service=Service(), options=options)

driver.get("https://groceries.asda.com")
wb = load_workbook("previous_prices.xlsx")
ws = wb.active

new_wb = Workbook()
new_ws = new_wb.active
new_ws.append(["Product", "Old Price", "New Price", "Change (%)"])

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
bold_font = Font(bold=True)

for row in ws.iter_rows(min_row=2, values_only=True):
    product, old = row
    driver.get("https://groceries.asda.com")
    try:
        search_box = driver.find_element(By.ID, "search-term-input")
        search_box.clear()
        search_box.send_keys(product)
        search_box.submit()
        time.sleep(3)

        price_element = driver.find_element(By.CSS_SELECTOR, '[data-testid="product-tile"] [data-testid="price"]')
        price_text = price_element.text.replace("£", "").replace(",", "").strip()
        new = float(price_text)
        change = round(((new - old) / old) * 100, 2)

        new_row = [product, old, new, change]
        new_ws.append(new_row)
        last_row = new_ws.max_row

        if change > 10:
            for col in range(1, 5):
                new_ws.cell(row=last_row, column=col).fill = green_fill
        elif change < -10:
            for col in range(1, 5):
                new_ws.cell(row=last_row, column=col).fill = red_fill
        if abs(change) > 20:
            for col in range(1, 5):
                new_ws.cell(row=last_row, column=col).font = bold_font

        print(f"✅ {product} - Old: {old}, New: {new}, Change: {change}%")
    except Exception as e:
        print(f"❌ {product} için veri alınamadı: {e}")

driver.quit()
new_wb.save("price_changes.xlsx")
print("📄 İşlem tamamlandı. price_changes.xlsx kaydedildi.")
